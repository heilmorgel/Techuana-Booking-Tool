"""Parse TECHUANA Gästeblatt Excel templates (.xlsx / .xltx)."""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.countries import COUNTRIES
from app.schemas import GaesteblattImportDraft, GaesteblattPersonDraft

_PERSON_START_ROW = 15
_PERSON_END_ROW = 208


def _norm_name(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date (Windows epoch 1899-12-30)
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return None
    text = str(value).strip()
    if not text or text.startswith("="):
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _build_nationality_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for item in COUNTRIES:
        code = item["code"].upper()
        lookup[code.lower()] = code
        lookup[_norm_name(item["name"])] = code
    # Common aliases from Meldeschein free text
    aliases = {
        "oesterreich": "AT",
        "österreich": "AT",
        "austria": "AT",
        "deutschland": "DE",
        "germany": "DE",
        "schweiz": "CH",
        "switzerland": "CH",
        "italien": "IT",
        "italy": "IT",
        "niederlande": "NL",
        "holland": "NL",
        "grossbritannien": "GB",
        "großbritannien": "GB",
        "vereinigtes koenigreich": "GB",
        "vereinigtes königreich": "GB",
        "uk": "GB",
        "usa": "US",
        "vereinigte staaten": "US",
        "tschechische republik": "CZ",
        "tschechien": "CZ",
        "slowakei": "SK",
        "slowenien": "SI",
        "kroatien": "HR",
        "ungarn": "HU",
        "polen": "PL",
        "frankreich": "FR",
        "spanien": "ES",
        "portugal": "PT",
        "belgien": "BE",
        "luxemburg": "LU",
        "daenemark": "DK",
        "dänemark": "DK",
        "schweden": "SE",
        "norwegen": "NO",
        "finnland": "FI",
        "rumaenien": "RO",
        "rumänien": "RO",
        "bulgarien": "BG",
        "griechenland": "GR",
        "tuerkei": "TR",
        "türkei": "TR",
        "ukraine": "UA",
        "serbien": "RS",
        "bosnien": "BA",
        "bosnien und herzegowina": "BA",
        "montenegro": "ME",
        "nordmazedonien": "MK",
        "mazedonien": "MK",
        "albanien": "AL",
        "kosovo": "XK",
        "irland": "IE",
        "island": "IS",
        "liechtenstein": "LI",
        "neuseeland": "NZ",
        "australien": "AU",
        "kanada": "CA",
    }
    for key, code in aliases.items():
        lookup[_norm_name(key)] = code
    return lookup


_NATIONALITY_LOOKUP = _build_nationality_lookup()


def map_nationality(raw: Any) -> str:
    text = _cell_str(raw)
    if not text:
        return "AT"
    if len(text) == 2 and text.isalpha():
        code = text.upper()
        if code in {c["code"] for c in COUNTRIES}:
            return code
    mapped = _NATIONALITY_LOOKUP.get(_norm_name(text))
    if mapped:
        return mapped
    return "XX"


def _join_name(last: str, first: str) -> str:
    parts = [p for p in (last.strip(), first.strip()) if p]
    return " ".join(parts)


def _format_leader_block(
    last: str,
    first: str,
    birth: date | None,
    travel_document: str,
    nationality: str,
    street: str,
    city: str,
    country: str,
) -> str:
    lines: list[str] = []
    name = _join_name(last, first)
    if name:
        lines.append(name)
    if birth:
        lines.append(f"Geburtsdatum: {birth.strftime('%d.%m.%Y')}")
    if nationality:
        lines.append(f"Staatsangehörigkeit: {nationality}")
    if travel_document:
        lines.append(f"Reisedokument: {travel_document}")
    if street:
        lines.append(street)
    if city:
        lines.append(city)
    if country:
        lines.append(country)
    return "\n".join(lines)


def _find_gaesteblatt_sheet(workbook):
    def normalize(name: str) -> str:
        text = unicodedata.normalize("NFKD", name or "")
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        return text.upper().replace("Ä", "A").replace("Ö", "O").replace("Ü", "U").replace("ß", "SS")

    for name in workbook.sheetnames:
        if "STEBLATT" in normalize(name):
            return workbook[name]
    raise ValueError("Arbeitsblatt 'Gästeblatt' nicht gefunden")


def _read_merged(ws, row: int, col: int) -> Any:
    """Return value for a cell, resolving merged ranges to the top-left cell."""
    cell = ws.cell(row, col)
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col).value
    return cell.value


def parse_gaesteblatt_bytes(content: bytes) -> GaesteblattImportDraft:
    # #region agent log
    import json as _json
    import time as _time
    from pathlib import Path as _Path

    _log_path = _Path(r"d:\Coding\Techuana_Homeassistant\debug-ad5dd0.log")

    def _dbg(hyp: str, loc: str, msg: str, data: dict) -> None:
        try:
            with _log_path.open("a", encoding="utf-8") as _f:
                _f.write(
                    _json.dumps(
                        {
                            "sessionId": "ad5dd0",
                            "runId": "repro",
                            "hypothesisId": hyp,
                            "location": loc,
                            "message": msg,
                            "data": data,
                            "timestamp": int(_time.perf_counter() * 1000)
                            if False
                            else int(_time.time() * 1000),
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )
        except Exception:
            pass

    # #endregion
    if not content:
        raise ValueError("Leere Datei")
    t0 = _time.perf_counter()
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Excel-Datei konnte nicht gelesen werden: {exc}") from exc
    _dbg("A", "gaesteblatt.py:load", "workbook loaded", {"seconds": round(_time.perf_counter() - t0, 3)})

    try:
        ws = _find_gaesteblatt_sheet(workbook)
    except ValueError:
        workbook.close()
        raise

    warnings: list[str] = []
    try:
        t1 = _time.perf_counter()
        merged_n = len(ws.merged_cells.ranges)
        group_name = _cell_str(_read_merged(ws, 8, 2))
        start_date = _parse_date(_read_merged(ws, 12, 2))
        end_date = _parse_date(_read_merged(ws, 13, 2))
        if start_date and end_date and start_date >= end_date:
            warnings.append("Anreise muss vor Abreise liegen — Daten bitte prüfen.")

        leader_last = _cell_str(_read_merged(ws, 4, 7))
        leader_first = _cell_str(_read_merged(ws, 5, 7))
        leader_birth = _parse_date(_read_merged(ws, 6, 7))
        leader_doc = _cell_str(_read_merged(ws, 7, 7))
        leader_nat_raw = _read_merged(ws, 8, 7)
        leader_nat = map_nationality(leader_nat_raw) if _cell_str(leader_nat_raw) else ""
        leader_street = _cell_str(_read_merged(ws, 9, 7))
        leader_city = _cell_str(_read_merged(ws, 10, 7))
        leader_country = _cell_str(_read_merged(ws, 11, 7))

        group_leader = _format_leader_block(
            leader_last,
            leader_first,
            leader_birth,
            leader_doc,
            _cell_str(leader_nat_raw) or leader_nat,
            leader_street,
            leader_city,
            leader_country,
        )

        persons: list[GaesteblattPersonDraft] = []
        empty_streak = 0
        rows_scanned = 0
        for row in range(_PERSON_START_ROW, _PERSON_END_ROW + 1):
            rows_scanned += 1
            last = _cell_str(_read_merged(ws, row, 2))
            if not last:
                empty_streak += 1
                continue
            empty_streak = 0
            first = _cell_str(_read_merged(ws, row, 3))
            birth = _parse_date(_read_merged(ws, row, 4))
            nationality = map_nationality(_read_merged(ws, row, 5))
            travel_document = _cell_str(_read_merged(ws, row, 6))
            person_start = _parse_date(_read_merged(ws, row, 8)) or start_date
            person_end = _parse_date(_read_merged(ws, row, 9)) or end_date
            persons.append(
                GaesteblattPersonDraft(
                    name=_join_name(last, first),
                    birth_date=birth,
                    nationality=nationality,
                    travel_document=travel_document,
                    start_date=person_start,
                    end_date=person_end,
                )
            )

        _dbg(
            "C",
            "gaesteblatt.py:persons",
            "person loop done",
            {
                "seconds": round(_time.perf_counter() - t1, 3),
                "merged": merged_n,
                "rows_scanned": rows_scanned,
                "persons": len(persons),
            },
        )

        leader_name = _join_name(leader_last, leader_first)
        if leader_name:
            existing = {_norm_name(p.name) for p in persons}
            if _norm_name(leader_name) not in existing:
                persons.insert(
                    0,
                    GaesteblattPersonDraft(
                        name=leader_name,
                        birth_date=leader_birth,
                        nationality=leader_nat or map_nationality(leader_country) or "AT",
                        travel_document=leader_doc,
                        start_date=start_date,
                        end_date=end_date,
                    ),
                )
                if not leader_birth:
                    warnings.append(
                        "Gruppenleiter wurde zur Personenliste hinzugefügt — Geburtsdatum fehlt."
                    )
                else:
                    warnings.append("Gruppenleiter wurde zur Personenliste hinzugefügt.")

        if not group_name:
            warnings.append("Gruppenname fehlt im Gästeblatt.")
        if not start_date or not end_date:
            warnings.append("Anreise/Abreise fehlen oder sind ungültig.")

        return GaesteblattImportDraft(
            group_name=group_name,
            start_date=start_date,
            end_date=end_date,
            group_leader=group_leader,
            persons=persons,
            warnings=warnings,
        )
    finally:
        workbook.close()
